# 创建扣分总表， 后面自己加了一些数据，这个数据录入还是要Excel来，这样搞没有数据太慢了
import openpyxl as op  # pip install openpyxl

# 接触尝试
# wb = op.Workbook()
# wb.create_sheet('demo')
# w = wb['demo']
# for i in range(10):
#     for i in range(10):
#         w.cell(row=i+1, column=i+1, value=i*i)
# wb.save('demo.xlsx')
# demo succeed

# 建表
wb = op.Workbook()
ws = wb.active
ws.title = '个人扣分'
# 确定表格信息
names = ['张三', '李四', '王五']
class_ = [1, 2, 5]
dormitory = [301, 520, 314]
num = [1, 7, 7]
mark = [-1, -0.5, -3]

# 写入信息
# 写入失败，以下是失败尝试
# rss = ws['A1':'E1']
# for rs in rss:
#     print(rs)
#     for r in rs:
#         print(r)
#         r.value
#
# ws.append(headers)

headers = ['姓名', '班级', '宿舍', '床号', '扣分']
for i in range(5):
    ws.cell(row=1, column=i+1, value=headers[i])
for i in range(1, 4):
    ws.cell(row=i + 1, column=1, value=names[i - 1])
    ws.cell(row=i + 1, column=2, value=class_[i - 1])
    ws.cell(row=i + 1, column=3, value=dormitory[i - 1])
    ws.cell(row=i + 1, column=4, value=num[i - 1])
    ws.cell(row=i + 1, column=5, value=mark[i - 1])

# 读取
ranges = ws['A1':'E4']
for i in ranges:
    li = []
    for j in i:
        li.append(j.value)
    print(li)

# 保存
wb.save('source.xlsx')
