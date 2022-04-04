import openpyxl
import sys
import shelve

sh = shelve.open('date')
if sh.keys():
    print(sh['date'])


# 基本配置数据
today_file = input("输入扣分数据日期：")
fps = [f'./data/{today_file}.xlsx']
koufen = input('扣分列：')
input('请确保关闭待操作文件，回车')
file_name = sys.argv[1]

sh['date'] = [today_file, koufen]
sh.close()

# *** 一、加载扣分原始数据

# 用于解析单条数据为多人扣分的情况，传入的参数为单条数据拆分后的列表
def parse_data(dt):
    people_num = len(dt) - 2
    for i in range(people_num):
        dormitory.append(int(dt[0]))
        beds.append(int(dt[i + 1]))
        scores.append(float(dt[-1]) / people_num)


# 将扣分数据按顺序排入列表，输入时也要按顺序
def load_data(today_file):
    wb = openpyxl.load_workbook(today_file)
    ws = wb.active

    for cell in ws['A']:
        datas.append(cell.value)
    wb.close()


datas = []
for fp in fps:
    load_data(fp)

dormitory = []
beds = []
scores = []
for data in datas:
    try:  # 防止输入有空白
        dt = data.split('/')
        parse_data(dt)
    except AttributeError:
        pass
print(dormitory, beds, scores)


# *** 二、扣分主体操作

# 主体扣分函数
def exec_dorm(ws, dorm):
    if dorm.value in dormitory:
        # 1.确定该宿舍范围
        min_row = dorm.row
        '''  用while写可以保证随意宿舍数量，切保证max_row正确'''
        i = 0
        while True:
            i += 1
            temp = ws.cell(column=2, row=min_row + i)
            if temp.value is not None and temp.value != dorm.value:
                break
            if temp.value is None and ws.cell(column=3, row=min_row+i).value is None:
                break
        max_row = min_row + i-1

        # 2.筛选同宿舍index，一次性操作完
        index = []
        for i in range(len(dormitory)):
            if dormitory[i] == dorm.value:
                index.append(i)

        # 3.进行扣分操作
        del_index = []  # 扣完分之后要删除的数据的索引
        for row in range(min_row, max_row+1):
            temp = ws.cell(column=3, row=row)
            for i in index:
                if temp.value == beds[i]:  # 找到目标，可以扣分（koufen是顶上的全局变量，扣分那列字母）
                    target = ws[f'{koufen}{temp.row}']
                    if target.value is not None:
                        target.value += scores[i]
                    else:
                        target.value = scores[i]
                    del_index.append(i)  # 可删除索引

        # 4.删除已扣分数据
        '''备注：之前列表从小到大排序，每次要修改del_index减1，现在从大到小排序，前面索引不变，速度更快'''
        del_index.sort(reverse=True)  # 因为添加时散乱，必须从大到小排序，从前往后逐一删除，避免破坏数据 （另：该方法直接排序列表，返回None）
        while del_index:  # 列表元素一直在减少，索引一直在改变，所以要小心
            del dormitory[del_index[0]]
            del beds[del_index[0]]
            del scores[del_index[0]]
            del del_index[0]


# 循环操作录入分数
wb = openpyxl.load_workbook(file_name)
for class_num in range(1, 21):
    worksheet = wb[f'{class_num}班']
    for dorm in worksheet['B']:
        exec_dorm(worksheet, dorm)
    print(f'{class_num}班:')
    print(dormitory, beds, scores, '\n')

wb.save(file_name)
wb.close()
input('done')
