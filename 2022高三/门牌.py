# 为了方便写，从门牌第二列开始，第一列请手动加上
import openpyxl

wb = openpyxl.load_workbook('高三男生个人扣分.xlsx')
ws = wb['Sheet1']
print('loaded...')
menpai = openpyxl.load_workbook('门牌8人间.xlsx')
sheet = menpai['Sheet1']
print('loaded...')

data = []
for i in range(4, 605, 9):
    li = []
    for row in range(i, i+8):
        name = ws.cell(row=row, column=1).value
        clas = ws.cell(row=row, column=4).value
        if clas:
            clas = '高三' + clas
        else:
            clas = None
        li.append((name, clas))
    data.append(tuple(li))
wb.close()
print('loaded data...')

count = 1
start = 1
for room in data:
    i = 0
    for name, clas in room:
        row = start + i
        if count == 1:
            sheet.cell(row=row, column=1, value=name)
            sheet.cell(row=row, column=2, value=clas)
        elif count == 2:
            sheet.cell(row=row, column=4, value=name)
            sheet.cell(row=row, column=5, value=clas)
        i += 1
    if count == 1:
        print(start, count, 'done', end='   ')
        count = 2
    elif count == 2:
        print(start, count, 'done')
        count = 1
        start += 9

menpai.save('门牌副本.xlsx')
print('done...')
