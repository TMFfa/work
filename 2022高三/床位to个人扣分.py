import openpyxl
import re

wb = openpyxl.load_workbook('C座高三男.xlsx')
ws = wb['Table 1']
print('loaded...')

gerenkoufen = openpyxl.load_workbook('高三男生个人扣分模板.xlsx')
sheet = gerenkoufen['Sheet1']
print('loaded...')

# 加载床位信息
li = []
for x in range(1, 68):
    room = ws.cell(row=x, column=1).value
    names = []
    for i in range(2, 10):
        names.append(ws.cell(row=x, column=i).value)
    classes = ws.cell(row=x, column=11).value
    li.append((room, names, classes))
    print('data loading...', x)
print('\ndata loaded...')
print('-'*20)
wb.close()

# 将床位信息录入到个人扣分表
for num, data in enumerate(li):
    start = (9 * num) +4
    print(start, data)
    chuanghao = 1
    for i in range(start, start+8):
        sheet.cell(row=i, column=2, value=data[0])
        sheet.cell(row=i, column=3, value=chuanghao)
        chuanghao += 1
    for i in range(8):
        row = start + i
        name = data[1][i]
        if name is None:
            continue
        clas = re.findall('\d+', name)
        if clas:
            clas = clas[0] + '班'
        else:
            clas = data[2]
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=4, value=clas)
    print('done')

gerenkoufen.save('副本.xlsx')
print('saved...')
