import openpyxl
from openpyxl.styles import Border, Side, colors, Alignment  # 用于格式设置

# 常量命名空间：
week = '第二周'
week_name = f'高三男生{week}周分.xlsx'

# 开始和结束列，就是要扣分的那几列，要扣哪几天的就输那几天列的英文序号（要大写）
start_column = 'J'
end_column = 'N'




all_wb = openpyxl.load_workbook('#高三男生个人扣分.xlsx')
all_ws = all_wb.active

week_wb = openpyxl.Workbook()  # 这是周分表  !!!!!新建的表不要命名，不要命名！否则不能进行常规操作！！！！

# 在这里分析起始和结束列的数字列号（其实openpyxl有内置的函数）
start_cell = all_ws[start_column + '1']
end_cell = all_ws[end_column + '1']
start_column = start_cell.column
end_column = end_cell.column
print(f'起始列：{start_cell.column}, 终止列{end_cell.column}')  # 它返回的会是数字列号





for class_ in range(1, 21):
# class_ = 1  # temp 临时使用

    # 创建并载入单个班级的表格
    sheet_name = f'{class_}班'
    print(sheet_name)
    week_wb.create_sheet(sheet_name)
    week_ws = week_wb[sheet_name]



    # ### 主体复制部分



    """
    该内容以制成函数在下面使用
    # 先复制第2行的表头，第一行的就不复制了
    # 注意这里是初版，有bug，周分表的列是固定的，不和个人扣分表一起
    for i in range(1, 5):
        value = all_ws.cell(row=2, column=i).value
        week_ws.cell(row=1, column=i, value=value)

    for i in range(start_column, end_column+1):
        value = all_ws.cell(row=2, column=i).value
        week_ws.cell(row=1, column=i, value=value)
    """

    # 这里将上面的复制写成函数，便于阅读
    def copy4(row1, row2):
        """
        这是前四列的复制
        row1:个人扣分表里的行列数
        row2:周分表里的行列数
        """
        for i in range(1, 5):
            value = all_ws.cell(row=row1, column=i).value
            week_ws.cell(row=row2, column=i, value=value)

    def copy_data(row1, row2, start_column, end_column):
        """
        这是扣分数据部分的复制，列数是变量，start_column和end_column是开头指定的
        row1,row2同上一个函数
        """
        column = 5  # 这是周分表的列，每次要递增
        for i in range(start_column, end_column+1):
            value = all_ws.cell(row=row1, column=i).value
            week_ws.cell(row=row2, column=column, value=value)
            column += 1



    # ######表头复制部分


    # 先复制第2行的表头，第一行的就不复制了
    copy4(row1=2, row2=1)

    # 这里用复制数据的函数来复制表格右边的表头（因为它不确定，是变量）
    # copy_data(row1=2, row2=1, start_column=start_column, end_column=end_column+1)
    # 这个表头是日期，比较特殊，直接复制的话会进行计算，所以还是不用函数

    column = 5  # 这是周分表的列，每次要递增
    for i in range(start_column, end_column+1):
        value = all_ws.cell(row=2, column=i).value
        week_ws.cell(row=1, column=column, value=value)
        week_ws.cell(row=1, column=column).number_format = 'm"月"d"日"'  # 这样才能显示日期!!!!
        column += 1




    # 数据复制部分


    # 表头搞定，下面将每个人的扣分复制过去
    week_ws_row = 2  # 记录周分表在哪一行，相当于打字的光标，一行行往下复制
    for cla in all_ws['D']:  # 遍历班级列，筛选对应班级。
        if cla.value == class_:
            row = cla.row
            copy4(row1=row, row2=week_ws_row)
            copy_data(row1=row, row2=week_ws_row, start_column=start_column, end_column=end_column)
            week_ws_row += 1


    # 下面给每行添加一个总分统计
    sum_column = week_ws.max_column + 1  # 扣分数据最大行
    week_ws.cell(row=1, column=sum_column, value='总分')
    column_letter = week_ws.cell(row=1, column=sum_column-1).column_letter  # 扣分数据最大列
    sum_letter = week_ws.cell(row=1, column=sum_column).column_letter  # 计算总分求和的那一列
    for row in range(2, week_ws.max_row+1):
        week_ws.cell(row=row, column=sum_column, value=f'=SUM(E{row}:{column_letter}{row})')
    # print(row)  # 可以继续使用for循环的row
    week_ws.cell(row=row+1, column=sum_column, value=f'=SUM({sum_letter}2:{sum_letter}{row})')
    week_ws.cell(row=row+1, column=1, value='总分')

    # 给总分行设置格式(合并单元格并居中)
    week_ws.merge_cells(f'A{row+1}:{column_letter}{row+1}')
    week_ws.cell(row=row+1, column=1).alignment = Alignment(horizontal='center', vertical='center')


    # 为所有单元格设置框线
    border = Border(left=Side(style='thin', color=colors.BLACK),
                    right=Side(style='thin', color=colors.BLACK),
                    top=Side(style='thin', color=colors.BLACK),
                    bottom=Side(style='thin', color=colors.BLACK),
    )
    for j in week_ws[week_ws.dimensions]:  # 或for row in week_ws.rows:然后再遍历一遍
        for k in j:
            k.border = border


week_wb.save(f'周分/{week_name}')
all_wb.close()
week_wb.close()
